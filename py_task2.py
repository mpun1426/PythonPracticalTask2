import openpyxl
import requests
import cchardet
import time
from bs4 import BeautifulSoup

url = 'https://www.i-parts.co.jp/used/index.asp?rs=0&pcm=7'
html = requests.get(url)
soup = BeautifulSoup(html.content, 'html.parser')
product_links = []
num = 1

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "スクレイピング課題2"

sheet["A1"].value = "商品名"
sheet["B1"].value = "販売価格（税込）"
sheet["C1"].value = "取扱店舗"
sheet["D1"].value = "商品コード"
sheet["E1"].value = "適合車種名称"
sheet.column_dimensions['A'].width = 45
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 13
sheet.column_dimensions['E'].width = 55

result = soup.find(class_="result-wrap")
items = result.find_all('td', class_="fst")

for item in items:
    product_link = item.find('a').attrs['href']
    product_links.append(product_link)

for product_page in product_links:
    product_html = requests.get('https://www.i-parts.co.jp' + product_page)
    product_html.encoding = cchardet.detect(html.content)["encoding"]
    product_soup = BeautifulSoup(product_html.text, "html.parser")
    product_contents = product_soup.select("#main")

    for product_info in product_contents:
        title = product_info.find('h2').text
        price = product_info.find_all('td')[0].text.rstrip("（税込）")
        shop = product_info.find_all('td')[2].find('a').text
        code = product_info.find_all('td')[4].text
        compatibility = product_info.find_all('td')[6].text
        sheet.cell(row = num + 1, column = 1, value = title)
        sheet.cell(row = num + 1, column = 2, value = price)
        sheet.cell(row = num + 1, column = 3, value = shop)
        sheet.cell(row = num + 1, column = 4, value = code)
        sheet.cell(row = num + 1, column = 5, value = compatibility)
        num += 1

    time.sleep(1)

wb.save("py_task2.xlsx")
wb.close()
